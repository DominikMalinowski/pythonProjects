import openpyxl

def enter_coordinates_manually():
    headers = ['X1', 'Y1', 'X2', 'Y2', 'Distance', 'Azimuth']
    coordinates = ['X1', 'Y1', 'X2', 'Y2']
    output_file_path = './modules/manually_provided_coordinates.xlsx'
    
    wb2 = openpyxl.Workbook()
    active_sheet = wb2.active

    for i in range(1,7):
        active_sheet.cell(row = 1, column = i).value = headers[i-1]

    for n in coordinates:
        print(f'Please provide coordinate for {n}')
        coordinate = float(input())
        active_sheet.cell(row = 2, column = (coordinates.index(n)+1)).value = coordinate

    wb2.save(output_file_path)
    return output_file_path
