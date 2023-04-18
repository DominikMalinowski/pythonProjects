import openpyxl
import math 
import time 
from modules import define_quarter as dq
from modules import input_method_selection as ims
from modules import enter_coordinates_manually as ecm

#TODO: 
# displaying appropriate message if user provide something different thant one of avaliable option 

selected_input_method = ims.input_method_selection()

def calculation(workbook):

    wb = openpyxl.load_workbook(workbook)
    active_sheet = wb.active
    active_sheet['E1'] = 'Distance'
    active_sheet['F1'] = 'Azimuth'

    for i in range(2, active_sheet.max_row + 1):
        print(f'Calculating distance and azimuth for row number: {i}')
        
        X1 = float((active_sheet.cell(row = i, column = 1).value))
        X2 = float((active_sheet.cell(row = i, column = 3).value))

        Y1 = float((active_sheet.cell(row = i, column = 2).value))
        Y2 = float((active_sheet.cell(row = i, column = 4).value))

        delta_X = X2 - X1
        delta_Y = Y2 - Y1

        distance = round(math.sqrt((delta_X)**2 + (delta_Y**2)), 2)
        azimuth = dq.define_quarter(delta_X, delta_Y)
        azimuth = float(round(azimuth, 6))

        active_sheet.cell(row = i, column = 5).value = distance
        active_sheet.cell(row = i, column = 6).value = azimuth
        
        wb.save('Results.xlsx')
    print('Calculation complete - please check "Results.xlsx" file')
    time.sleep(2)

while True:
    if selected_input_method == '1':
        calculation(ecm.enter_coordinates_manually())
        break
    elif selected_input_method == '2':
        calculation('Input File Template.xlsx')
        break
    else: 
        print('Please select one of provided option')
        continue