import openpyxl
import math 
from modules import define_quarter as dq


def calculation(workbook):
    """
    Base function that calculate value of fi and decide how to use in calculation base on value for delta X and Y 
    """
    wb = openpyxl.load_workbook(workbook)
    active_sheet = wb.active

    for i in range(2, active_sheet.max_row + 1):
        print(f'Calculating distance and azimuth for row number: {i}')
        
        X1 = float((active_sheet.cell(row = i, column = 1).value))
        X2 = float((active_sheet.cell(row = i, column = 3).value))

        Y1 = float((active_sheet.cell(row = i, column = 2).value))
        Y2 = float((active_sheet.cell(row = i, column = 4).value))

        delta_X = X2 - X1
        delta_Y = Y2 - Y1

        fi = math.degrees(math.atan(abs(delta_Y/delta_X)))
        distance = round(math.sqrt((delta_X)**2+(delta_Y**2)),2)

        azimuth = float(round(dq.define_quarter(delta_X, delta_Y, fi),6))
        
        active_sheet.cell(row = 1, column = 5).value = 'Distance'
        active_sheet.cell(row = i, column = 5).value = distance

        active_sheet.cell(row = 1, column = 6).value = 'Azimuth'
        active_sheet.cell(row = i, column = 6).value = azimuth

        wb.save('excel_test_file_output.xlsx')
   
    print('Calculation complete')


calculation('excel_test_file.xlsx')