import openpyxl
from openpyxl import utils 

wb = openpyxl.load_workbook('example.xlsx')

# get all sheets name 
print(wb.get_sheet_names())

#saving specyfic sheet as variable
sheet_3 = wb.get_sheet_by_name('Sheet3')

# getting active sheet from the file 
active_sheet = wb.active
print(active_sheet.title)

# getting value from specyfic cell 
A1 = active_sheet['A1']

print(A1.coordinate)
print(A1.value)
print(A1.row)
print(A1.column)

for i in range (1, 8, 2):
    print(i, active_sheet.cell(row=i, column=2).value)

# getting data from specyfic row/column 
placeholder = active_sheet.cell(row = 1, column = 2).value

# getting max row/column number 
row_count = active_sheet.max_row
column_count = active_sheet.max_column

print(row_count, column_count)

# getting #n column name(letter) 
print(utils.get_column_letter(115))

# getting column # from name
print(utils.column_index_from_string('DK'))

# displaying all data in range 
for rowOfCellObjects in active_sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('Koniec')



