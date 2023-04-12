import openpyxl
import math 


def calculation(workbook):
    """
    Base function that calculate value of fi and decide how to use in calculation base on value for delta X and Y 
    """
    wb = openpyxl.load_workbook(workbook)
    active_sheet = wb.active
    results = {}
    for i in range(2, active_sheet.max_row + 1):
        print(i)
        azimuth_name = ('Azimuth for row ' + str(i))
        distance_name = ('Distance for row ' + str(i))
        
        X1 = float((active_sheet.cell(row = i, column = 1).value))
        X2 = float((active_sheet.cell(row = i, column = 3).value))

        Y1 = float((active_sheet.cell(row = i, column = 2).value))
        Y2 = float((active_sheet.cell(row = i, column = 4).value))

        delta_X = X2 - X1
        delta_Y = Y2 - Y1

        fi = math.degrees(math.atan(abs(delta_Y/delta_X)))
        distance = round(math.sqrt((delta_X)**2+(delta_Y**2)),2)

        if delta_X == 0:
            if delta_Y == 0:
                azimuth = 'Special case - same place'
            elif delta_Y > 0:
                azimuth = 90
            else:
                azimuth = 240 
        elif delta_X > 0:
            if delta_Y == 0:
                azimuth = 0
            elif delta_Y > 0:
                azimuth = fi 
            else:
                azimuth = 360- fi
        elif delta_X < 0:
            if delta_Y == 0:
                azimuth = 180
            elif delta_Y > 0:
                azimuth = 180 - fi 
            else:
                azimuth = 180 + fi

        azimuth = float(round(azimuth,6))
        
        active_sheet.cell(row = i, column = 5).value = distance
        active_sheet.cell(row = i, column = 6).value = azimuth

        wb.save('excel_test_file_output.xlsx')

        results.update({azimuth_name:float(round(azimuth,6))})
        results.update({distance_name: distance})
   
    return(results)
    
    
calculation('excel_test_file.xlsx')