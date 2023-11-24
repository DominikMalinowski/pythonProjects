# opening excel file 
import openpyxl
workbook = openpyxl.load_workbook('{file_name}')

# getting active spreadsheet 
import openpyxl
workbook = openpyxl.load_workbook('{file_name}')
active_sheet = workbook.get_active_sheet()

# getting spreadsheet by name 
import openpyxl
workbook = openpyxl.load_workbook('{file_name}')
sheet = workbook.get_sheet_by_name('{sheet_name}')

# getting cell and it's value from spreadsheet 
import openpyxl
workbook = openpyxl.load_workbook('{file_name}')
sheet = workbook.get_sheet_by_name('{sheet_name}')
cell = sheet['A1']
cell2 = sheet(row=1, column =2)
cell.value 
cell2.value

# getting coordinate and value from spreadsheet 
import openpyxl
workbook = openpyxl.load_workbook('{file_name}')
sheet = workbook.get_sheet_by_name('{sheet_name}')
cell = sheet['A1']
coordinate = cell.coordinate 
value = cell.value  

# changing spreadsheet name
import openpyxl
workbook = openpyxl.load_workbook('{file_name}')
sheet = workbook.get_sheet_by_name('{sheet_name}')
sheet.title = '{new title}'

# saving spreadsheet
import openpyxl
workbook = openpyxl.load_workbook('{file_name}')
sheet = workbook.get_sheet_by_name('{sheet_name}')
workbook.save('{new_file_name_with_extension}')

# change column letter and number 
import openpyxl
workbook = openpyxl.load_workbook('{file_name}')
workbook.get_column_letter(1)

# creating new spreadsheet 
import openpyxl
workbook = openpyxl.load_workbook('{file_name}')
workbook.create_sheet(index = 0, title = 'sheet_add')
workbook.create_sheet(index = 2, title = 'sheet_add2')

# removing spreadsheet
import openpyxl
workbook = openpyxl.load_workbook('{file_name}')
workbook.remove_sheet('{sheet_to_remove}')

# populating cell 
import openpyxl
workbook = openpyxl.load_workbook('{file_name}')
sheet = workbook.get_sheet_by_name('{sheet_name}')
sheet['A1'] = 'Placeholder'

