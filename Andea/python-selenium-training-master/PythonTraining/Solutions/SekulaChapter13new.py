from random import randint
import openpyxl
import pyinputplus as p
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# method set values in table headers
def SetTableHeader(numOfCell):
    fontObj = Font(bold=True, size=15)
    for i in range(2, numOfCell+2):
        tableSheet["A"+str(i)] = i-1
        tableSheet["A" + str(i)].font=fontObj
        tableSheet[get_column_letter(i) + "1"] = i-1
        tableSheet[get_column_letter(i) + "1"].font = fontObj
    tableSheet.freeze_panes = "B2"
    wb.save("Test.xlsx")

#method fill table by proper values
def FillTable(numofCell):
    for i in range(2, numOfCell+2):
        for j in range(2, numOfCell+2):
            tableSheet[get_column_letter(j) + str(i)] = int(tableSheet["A" + str(i)].value) * int(tableSheet[get_column_letter(j) + "1"].value)
    wb.save("Test.xlsx")

# method rand two cells from table and sum values
def SumTwoValues(counter):
    for x in range(counter):
        firstValue = get_column_letter(randint(1, numOfCell)) + str(randint(2, numOfCell+1))
        secondValue = get_column_letter(randint(1, numOfCell)) + str(randint(2, numOfCell+1))
        result = int(tableSheet[firstValue].value) * int(tableSheet[secondValue].value)
        resultSheet['A'+str(x+1)] = result
    wb.save("Test.xlsx")



wb = openpyxl.Workbook()
tableSheet = wb.active
tableSheet.title = "Table"
resultSheet = wb.create_sheet("Results")
numOfCell = 100
numOfResults = p.inputInt("Enter number of results: ")

SetTableHeader(numOfCell)
FillTable(numOfCell)
SumTwoValues(numOfResults)